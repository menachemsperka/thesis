import torch
import torch.nn as nn
import torch.optim as optim
import numpy as np
from torch.utils.data import DataLoader, Dataset
from transformers import AutoTokenizer, AutoModel
import pandas as pd
import chardet
from tqdm import tqdm
from sklearn.metrics import confusion_matrix, classification_report
import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows
import os
import random

_internal_model = os.path.join(os.path.dirname(os.path.dirname(__file__)), 'models', 'dictabert')
DEFAULT_MODEL_NAME = os.environ.get(
    'THESIS_MODEL_NAME',
    _internal_model if os.path.exists(os.path.join(_internal_model, 'config.json')) else 'dicta-il/dictabert'
)


def _local_only_enabled():
    return os.environ.get('THESIS_MODEL_LOCAL_ONLY', '0').strip().lower() in ('1', 'true', 'yes', 'on')

# Define global training parameters
LEARNING_RATE = 5e-5
NUM_EPOCHS = 3
BATCH_SIZE = 16
LAMBDA_PARAM = 100.0
MARGIN = 1.0

class NERDataset(Dataset):
    """Custom dataset for the reformulated NER task."""
    def __init__(self, sentences, entity_labels, begin_labels):
        self.sentences = sentences
        self.entity_labels = entity_labels
        self.begin_labels = begin_labels

    def __len__(self):
        return len(self.sentences)

    def __getitem__(self, idx):
        return {
            'sentence': self.sentences[idx],
            'entity_labels': self.entity_labels[idx],
            'begin_labels': self.begin_labels[idx]
        }

class NERModelAUC2T(nn.Module):
    """Model for AUC-2T NER with binary reformulation."""
    def __init__(self, model_name=DEFAULT_MODEL_NAME, num_classes=2):
        super(NERModelAUC2T, self).__init__()
        local_only = os.environ.get('THESIS_MODEL_LOCAL_ONLY', '0').strip().lower() in ('1', 'true', 'yes', 'on')
        self.base_model = AutoModel.from_pretrained(model_name, local_files_only=local_only)
        hidden_size = self.base_model.config.hidden_size
        
        # Binary classifiers for entity token and beginning token
        self.entity_classifier = nn.Linear(hidden_size, 1)
        self.begin_classifier = nn.Linear(hidden_size, 1)
        
    def get_shared_embeddings(self, input_ids, attention_mask=None):
        """Get embeddings from the base model."""
        outputs = self.base_model(
            input_ids=input_ids,
            attention_mask=attention_mask,
            return_dict=True
        )
        return outputs.last_hidden_state
    
    def entity_token_classifier(self, embeddings):
        """Apply entity token classifier to embeddings."""
        return self.entity_classifier(embeddings).squeeze(-1)
    
    def beginning_token_classifier(self, embeddings):
        """Apply beginning token classifier to embeddings."""
        return self.begin_classifier(embeddings).squeeze(-1)
    
    def forward(self, input_ids, attention_mask=None):
        embeddings = self.get_shared_embeddings(input_ids, attention_mask)
        entity_scores = self.entity_token_classifier(embeddings)
        begin_scores = self.beginning_token_classifier(embeddings)
        return entity_scores, begin_scores

def calculate_DAM_loss(scores, labels, margin=1.0, mask=None):
    """
    Calculate Deep AUC Margin (DAM) loss.
    
    Args:
        scores: Model prediction scores (batch_size, seq_length)
        labels: Ground truth labels (1 for positive, -1 for negative) (batch_size, seq_length)
        margin: Margin parameter for the loss
        mask: Optional mask for padding tokens (1 for valid tokens, 0 for padding)
    
    Returns:
        AUCM loss value
    """
    device = scores.device
    
    # If no mask provided, create one assuming all tokens are valid
    if mask is None:
        mask = torch.ones_like(scores, device=device)
    
    # Ensure mask and scores have the same shape
    if mask.shape != scores.shape:
        # Resize mask to match scores shape
        if len(mask.shape) == 2 and len(scores.shape) == 2:
            # If both are 2D but with different sizes, use the smaller length for each dimension
            batch_size = min(mask.shape[0], scores.shape[0])
            seq_len = min(mask.shape[1], scores.shape[1])
            new_mask = torch.zeros_like(scores, device=device)
            new_mask[:batch_size, :seq_len] = mask[:batch_size, :seq_len]
            mask = new_mask
        else:
            # For other dimension mismatches, create a new mask of ones
            mask = torch.ones_like(scores, device=device)
    
    # Flatten all tensors for easier processing
    flat_scores = scores.view(-1)
    flat_labels = labels.view(-1)
    flat_mask = mask.view(-1)
    
    # Only consider valid tokens (not padding)
    valid_indices = flat_mask.bool()
    if not torch.any(valid_indices):
        return torch.tensor(0.0, device=device, requires_grad=True)
    
    valid_scores = flat_scores[valid_indices]
    valid_labels = flat_labels[valid_indices]
    
    # Separate positive and negative examples
    pos_indices = (valid_labels == 1)
    neg_indices = (valid_labels == -1)
    
    if not torch.any(pos_indices) or not torch.any(neg_indices):
        # If there are no positive or negative examples, return zero loss
        return torch.tensor(0.0, device=device, requires_grad=True)
    
    pos_scores = valid_scores[pos_indices]
    neg_scores = valid_scores[neg_indices]
    
    # Number of positive and negative samples
    n_pos = pos_scores.size(0)
    n_neg = neg_scores.size(0)
    
    # Calculating means
    pos_mean = torch.mean(pos_scores)
    neg_mean = torch.mean(neg_scores)
    
    # Calculating variances
    pos_var = torch.var(pos_scores, unbiased=False) if n_pos > 1 else torch.tensor(0.0, device=device)
    neg_var = torch.var(neg_scores, unbiased=False) if n_neg > 1 else torch.tensor(0.0, device=device)
    
    # Variance regularization (keeps score distributions tight)
    var_loss = 0.5 * (pos_var + neg_var)
    
    # Squared margin loss (DAM formulation)
    # Using squared form instead of relu to maintain smooth gradients everywhere
    # and prevent gradient death once the margin is trivially achieved.
    margin_gap = neg_mean - pos_mean + margin
    margin_loss = margin_gap ** 2
    
    loss = var_loss + margin_loss
    
    return loss

def transform_bio_labels_to_binary(original_labels):
    """
    Transform original BIO labels to binary labels for entity and beginning tokens.
    
    Args:
        original_labels: List of original BIO labels for each token (string or numeric)
        
    Returns:
        entity_labels: Binary labels for the entity-token task (1 for entity, -1 for non-entity)
        begin_labels: Binary labels for the beginning-token task (1 for beginning, -1 for non-beginning)
    """
    entity_labels = []
    begin_labels = []
    
    for label in original_labels:
        # Handle numeric labels if needed
        if isinstance(label, (int, float)) or (isinstance(label, str) and label.isdigit()):
            # Based on your dataset: 10 appears to be "O", 1 appears to be "B-"
            label_num = int(label)
            # This mapping needs adjustment based on your actual numeric encoding
            if label_num == 10:  # Assuming 10 is "O"
                entity_labels.append(-1)  # Not an entity
                begin_labels.append(-1)   # Not a beginning
            elif label_num == 1:  # Assuming 1 is "B-something"
                entity_labels.append(1)   # Is an entity
                begin_labels.append(1)    # Is a beginning
            elif label_num > 1 and label_num != 10:  # Assuming other numbers are "I-something"
                entity_labels.append(1)   # Is an entity
                begin_labels.append(-1)   # Not a beginning
            else:
                # Default case
                entity_labels.append(-1)
                begin_labels.append(-1)
        else:
            # String labels - use normal BIO format handling
            # Entity token labeling (B- or I- are entities)
            if isinstance(label, str) and (label.startswith('B-') or label.startswith('I-')):
                entity_labels.append(1)  # Positive for entity
            else:
                entity_labels.append(-1)  # Negative for entity (O)
            
            # Beginning token labeling (only B- are beginnings)
            if isinstance(label, str) and label.startswith('B-'):
                begin_labels.append(1)   # Positive for beginning
            else:
                begin_labels.append(-1)  # Negative for beginning (I- or O)
            
    return entity_labels, begin_labels

def preprocess_data(data, tokenizer, connect_subwords=True):
    """
    Preprocess training data into binary classification format.
    
    Args:
        data: DataFrame containing sentences and their original BIO labels
        tokenizer: Tokenizer for the model
        connect_subwords: If True, merge subwords (starting with ##) into one word
        
    Returns:
        transformed_data: List of dictionaries with input_ids, attention_mask, entity_labels, begin_labels
    """
    # Check the actual column names in the dataframe
    columns = data.columns.tolist()
    
    # Determine which columns to use based on what's available
    sentence_col = 'Sentence #' if 'Sentence #' in columns else 'id'
    word_col = 'Word' if 'Word' in columns else 'token'
    
    # Determine tag column - prefer raw_tags if available, otherwise Tag or ner_tags
    if 'raw_tags' in columns:
        tag_col = 'raw_tags'
    elif 'Tag' in columns:
        tag_col = 'Tag'
    elif 'ner_tags' in columns:
        # If using numeric tags, we'll need to map them to string format
        # This assumes you have a mapping somewhere - otherwise use actual raw_tags
        print("Warning: Using numeric ner_tags, make sure they are properly mapped to BIO format")
        tag_col = 'ner_tags'
    else:
        raise ValueError("No suitable tag column found in the dataset")
    
    # Group by Sentence # to get sentences and their token labels
    sentences = []
    all_entity_labels = []
    all_begin_labels = []
    
    for sentence_id, group in data.groupby(sentence_col):
        words = group[word_col].tolist()
        original_labels = group[tag_col].tolist()
        
        # Handle subword merging if connect_subwords is True
        if connect_subwords:
            merged_words = []
            merged_labels = []
            current_word = ""
            current_label = None
            
            for word, label in zip(words, original_labels):
                word = str(word)  # Ensure word is a string
                if word.startswith("##"):
                    current_word += word[2:]  # Append subword without ##
                else:
                    if current_word:  # Add the previous merged word
                        merged_words.append(current_word)
                        merged_labels.append(current_label)
                    current_word = word
                    current_label = label
            
            # Add the last merged word
            if current_word:
                merged_words.append(current_word)
                merged_labels.append(current_label)
            
            words = merged_words
            original_labels = merged_labels
        
        # Filter out special tokens like [CLS], [SEP] if present
        filtered_words = []
        filtered_labels = []
        for word, label in zip(words, original_labels):
            if word not in ['[CLS]', '[SEP]', ''] and not pd.isna(word):
                filtered_words.append(word)
                filtered_labels.append(label)
        
        if not filtered_words:  # Skip empty sentences
            continue
            
        # Transform original BIO labels to binary labels
        entity_labels, begin_labels = transform_bio_labels_to_binary(filtered_labels)
        
        sentences.append(' '.join(filtered_words))
        all_entity_labels.append(entity_labels)
        all_begin_labels.append(begin_labels)
    
    return sentences, all_entity_labels, all_begin_labels

def train_auc_2t(
    data,
    model_name=DEFAULT_MODEL_NAME,
    learning_rate=LEARNING_RATE,  # Use global variable
    num_epochs=NUM_EPOCHS,  # Use global variable
    batch_size=BATCH_SIZE,  # Use global variable
    lambda_param=LAMBDA_PARAM,  # Use global variable
    margin=MARGIN,  # Use global variable
    device='cuda' if torch.cuda.is_available() else 'cpu'
):
    """
    Train a NER model using the AUC-2T method.
    
    Args:
        data: DataFrame containing NER data with columns: 'Sentence #', 'Word', 'Tag'
        model_name: Name of the pre-trained model to use
        learning_rate: Learning rate for the optimizer
        num_epochs: Number of training epochs
        batch_size: Batch size for training
        lambda_param: Trade-off parameter for combining the two AUC losses
        margin: Margin parameter for the Deep AUC Margin loss
        device: Device to use for training ('cuda' or 'cpu')
        
    Returns:
        model: Trained NER model
        history: Dictionary containing training metrics
    """
    local_only = os.environ.get('THESIS_MODEL_LOCAL_ONLY', '0').strip().lower() in ('1', 'true', 'yes', 'on')
    tokenizer = AutoTokenizer.from_pretrained(model_name, local_files_only=local_only)
    model = NERModelAUC2T(model_name=model_name).to(device)
    
    # Step 1: Data Preprocessing (Reformulation)
    print("Preprocessing data...")
    sentences, entity_labels, begin_labels = preprocess_data(data, tokenizer)
    
    # Create dataset and dataloader
    dataset = NERDataset(sentences, entity_labels, begin_labels)
    dataloader = DataLoader(dataset, batch_size=batch_size, shuffle=True, collate_fn=custom_collate_fn)
    
    # Step 2: Initialize Model and Optimizer
    optimizer = optim.Adam(model.parameters(), lr=learning_rate)
    
    # Step 3: Training Loop
    print(f"Starting training for {num_epochs} epochs...")
    history = {
        'entity_loss': [],
        'begin_loss': [],
        'total_loss': []
    }
    
    for epoch in range(num_epochs):
        model.train()
        epoch_entity_loss = 0
        epoch_begin_loss = 0
        epoch_total_loss = 0
        batch_count = 0
        
        for batch in dataloader:
            sentences = batch['sentence']
            entity_labels_batch = batch['entity_labels']
            begin_labels_batch = batch['begin_labels']
            
            # Split sentences back into word lists for proper subword alignment
            word_lists = [s.split() for s in sentences]
            tokenized = tokenizer(
                word_lists,
                is_split_into_words=True,
                padding=True,
                truncation=True,
                return_tensors='pt',
                return_special_tokens_mask=True
            )

            # Collect word_ids before moving to device (word_ids relies on internal _encodings)
            batch_word_ids = [tokenized.word_ids(batch_index=i) for i in range(len(sentences))]
            tokenized = tokenized.to(device)
            
            input_ids = tokenized['input_ids']
            attention_mask = tokenized['attention_mask']
            special_tokens_mask = tokenized['special_tokens_mask']
            
            batch_size, max_seq_len = input_ids.shape
            
            # Initialize label tensors with -1 (default negative label)
            entity_labels_tensor = torch.ones((batch_size, max_seq_len), device=device) * -1
            begin_labels_tensor = torch.ones((batch_size, max_seq_len), device=device) * -1
            
            # Create a mask for valid tokens (non-padding and non-special)
            valid_tokens_mask = (attention_mask == 1) & (special_tokens_mask == 0)
            valid_tokens_mask = valid_tokens_mask.float()
            
            # Align word-level labels with subword tokens using word_ids
            for i in range(len(sentences)):
                orig_entity_labels = entity_labels_batch[i]
                orig_begin_labels = begin_labels_batch[i]
                word_ids = batch_word_ids[i]
                prev_word_id = None
                word_idx = -1
                
                for j in range(max_seq_len):
                    wid = word_ids[j] if j < len(word_ids) else None
                    if wid is None:
                        continue  # Special token or padding
                    
                    is_first_subword = (wid != prev_word_id)
                    if is_first_subword:
                        word_idx += 1
                    prev_word_id = wid
                    
                    if word_idx < len(orig_entity_labels):
                        entity_labels_tensor[i][j] = orig_entity_labels[word_idx]
                        if is_first_subword:
                            begin_labels_tensor[i][j] = orig_begin_labels[word_idx]
                        else:
                            begin_labels_tensor[i][j] = -1  # Subword continuation: never a beginning
            
            # Forward Pass
            optimizer.zero_grad()
            
            # Get shared embeddings and apply classifiers
            shared_embeddings = model.get_shared_embeddings(input_ids, attention_mask)
            entity_scores = model.entity_token_classifier(shared_embeddings)
            begin_scores = model.beginning_token_classifier(shared_embeddings)
            
            # Calculate losses with masks
            loss_en = calculate_DAM_loss(entity_scores, entity_labels_tensor, margin, valid_tokens_mask)
            loss_be = calculate_DAM_loss(begin_scores, begin_labels_tensor, margin, valid_tokens_mask)
            
            # Combine losses
            total_loss = loss_en + lambda_param * loss_be
            
            # Backward pass
            total_loss.backward()
            optimizer.step()
            
            # Track losses
            epoch_entity_loss += loss_en.item()
            epoch_begin_loss += loss_be.item()
            epoch_total_loss += total_loss.item()
            batch_count += 1
        
        # Calculate average losses for the epoch
        avg_entity_loss = epoch_entity_loss / max(batch_count, 1)
        avg_begin_loss = epoch_begin_loss / max(batch_count, 1)
        avg_total_loss = epoch_total_loss / max(batch_count, 1)
        
        # Store metrics
        history['entity_loss'].append(avg_entity_loss)
        history['begin_loss'].append(avg_begin_loss)
        history['total_loss'].append(avg_total_loss)
        
        print(f"Epoch {epoch+1}/{num_epochs}, "
              f"Entity Loss: {avg_entity_loss:.4f}, "
              f"Begin Loss: {avg_begin_loss:.4f}, "
              f"Total Loss: {avg_total_loss:.4f}")
    
    print("Training completed!")
    
    return model, history

def predict_with_auc_2t(model, sentences, tokenizer, threshold=0.0, device='cuda' if torch.cuda.is_available() else 'cpu'):
    """
    Make predictions using the trained AUC-2T model.
    
    Args:
        model: Trained AUC-2T model
        sentences: List of sentences to predict
        tokenizer: Tokenizer for the model
        threshold: Decision threshold for binary classification
        device: Device to use for inference
        
    Returns:
        entity_predictions: Binary predictions for entity tokens
        begin_predictions: Binary predictions for beginning tokens
    """
    model.eval()
    entity_predictions = []
    begin_predictions = []
    
    with torch.no_grad():
        for sentence in sentences:
            # Split sentence back into word list for proper subword alignment
            words = sentence.split()
            tokenized = tokenizer(
                words,
                is_split_into_words=True,
                padding=True,
                truncation=True,
                return_tensors='pt',
            )

            # Collect word_ids before moving to device
            word_ids = tokenized.word_ids(batch_index=0)
            tokenized = tokenized.to(device)
            
            # Get predictions
            entity_scores, begin_scores = model(
                tokenized['input_ids'],
                tokenized['attention_mask']
            )
            
            # Apply threshold
            entity_preds = (entity_scores > threshold).int() * 2 - 1  # Convert to -1, 1
            begin_preds = (begin_scores > threshold).int() * 2 - 1   # Convert to -1, 1
            
            # Aggregate subword predictions to word level using word_ids
            word_entity_preds = []
            word_begin_preds = []
            prev_word_id = None
            
            for j, wid in enumerate(word_ids):
                if wid is None:
                    continue  # Skip special tokens and padding
                if wid != prev_word_id:
                    # First subword of new word - take its prediction
                    word_entity_preds.append(entity_preds[0][j].item())
                    word_begin_preds.append(begin_preds[0][j].item())
                    prev_word_id = wid
            
            entity_predictions.append(word_entity_preds)
            begin_predictions.append(word_begin_preds)
    
    return entity_predictions, begin_predictions

def convert_binary_to_bio(entity_preds, begin_preds, entity_types=None):
    """
    Convert binary predictions back to BIO format.
    
    Args:
        entity_preds: Binary predictions for entity tokens (-1 or 1)
        begin_preds: Binary predictions for beginning tokens (-1 or 1)
        entity_types: List of entity types (if None, uses generic "ENT" type)
        
    Returns:
        bio_tags: List of BIO tags
    """
    if entity_types is None:
        entity_types = ["ENT"]  # Default entity type
    
    bio_tags = []
    current_entity = False
    
    # Make sure we're working with Python lists, not numpy arrays
    if hasattr(entity_preds, 'tolist'):
        entity_preds = entity_preds.tolist()
    if hasattr(begin_preds, 'tolist'):
        begin_preds = begin_preds.tolist()
    
    for i in range(len(entity_preds)):
        is_entity = (entity_preds[i] == 1)
        is_begin = (begin_preds[i] == 1)
        
        if not is_entity:
            bio_tags.append("O")
            current_entity = False
        elif is_begin:
            # For simplicity, always use the first entity type
            bio_tags.append(f"B-{entity_types[0]}")
            current_entity = True
        elif current_entity:
            bio_tags.append(f"I-{entity_types[0]}")
        else:
            # This shouldn't happen in theory but we'll handle it by assuming it's a beginning
            bio_tags.append(f"B-{entity_types[0]}")
            current_entity = True
    
    return bio_tags

def evaluate_model(model, test_data, tokenizer, entity_types=None, device='cuda' if torch.cuda.is_available() else 'cpu'):
    """
    Evaluate the AUC-2T model on test data.
    
    Args:
        model: Trained AUC-2T model
        test_data: DataFrame containing test data
        tokenizer: Tokenizer for the model
        entity_types: List of entity types
        device: Device to use for inference
        
    Returns:
        metrics: Dictionary containing evaluation metrics
    """
    # Import here to avoid requiring seqeval if not evaluating
    try:
        from seqeval.metrics import classification_report, accuracy_score, precision_score, recall_score, f1_score
    except ImportError:
        print("Warning: seqeval not installed, metrics will be limited")
        # Define dummy functions that return 0
        def accuracy_score(*args, **kwargs): return 0
        def precision_score(*args, **kwargs): return 0
        def recall_score(*args, **kwargs): return 0
        def f1_score(*args, **kwargs): return 0
        def classification_report(*args, **kwargs): return {}
    
    # Preprocess test data
    sentences, true_entity_labels, true_begin_labels = preprocess_data(test_data, tokenizer)
    
    # Get predictions
    pred_entity_labels, pred_begin_labels = predict_with_auc_2t(model, sentences, tokenizer, device=device)
    
    # Convert binary predictions back to BIO format
    true_bio_tags = []
    pred_bio_tags = []
    
    for sentence_idx in range(len(sentences)):
        # Convert predicted binary labels to BIO tags
        try:
            sentence_pred_bio = convert_binary_to_bio(
                pred_entity_labels[sentence_idx], 
                pred_begin_labels[sentence_idx],
                entity_types
            )
            pred_bio_tags.append(sentence_pred_bio)
            
            # Get true BIO tags
            entity_is_positive = true_entity_labels[sentence_idx]
            begin_is_positive = true_begin_labels[sentence_idx]
            
            # Reconstruct BIO tags - make sure to handle possible length mismatch
            sentence_true_bio = []
            for e, b in zip(entity_is_positive, begin_is_positive):
                if e == -1:  # Not an entity
                    sentence_true_bio.append("O")
                elif b == 1:  # Beginning of entity
                    sentence_true_bio.append(f"B-{entity_types[0] if entity_types else 'ENT'}")
                else:  # Inside entity
                    sentence_true_bio.append(f"I-{entity_types[0] if entity_types else 'ENT'}")
            true_bio_tags.append(sentence_true_bio)
            
            # Ensure pred and true tags have the same length
            min_len = min(len(sentence_true_bio), len(sentence_pred_bio))
            true_bio_tags[-1] = true_bio_tags[-1][:min_len]
            pred_bio_tags[-1] = pred_bio_tags[-1][:min_len]
        
        except Exception as e:
            print(f"Error processing sentence {sentence_idx}: {e}")
            # Skip this sentence
            continue
    
    # Calculate metrics using seqeval
    try:
        report = classification_report(true_bio_tags, pred_bio_tags, output_dict=True)
    except Exception as e:
        print(f"Error calculating classification report: {e}")
        report = {}
    
    # Extract overall metrics
    try:
        metrics = {
            'accuracy': accuracy_score(true_bio_tags, pred_bio_tags),
            'precision': precision_score(true_bio_tags, pred_bio_tags),
            'recall': recall_score(true_bio_tags, pred_bio_tags),
            'f1': f1_score(true_bio_tags, pred_bio_tags),
            'report': report
        }
    except Exception as e:
        print(f"Error calculating metrics: {e}")
        metrics = {'error': str(e)}
    
    return metrics

def custom_collate_fn(batch):
    """
    Custom collate function to handle variable length sequences.
    
    Args:
        batch: List of samples from the dataset
        
    Returns:
        Batched data with appropriate padding
    """
    sentences = [item['sentence'] for item in batch]
    entity_labels = [item['entity_labels'] for item in batch]
    begin_labels = [item['begin_labels'] for item in batch]
    
    # We don't need to do anything special with the sentences
    # as they will be tokenized and padded by the tokenizer later
    
    # For the labels, we'll use a simple approach: return them as lists
    # They'll be properly aligned with tokenized outputs during training
    return {
        'sentence': sentences,
        'entity_labels': entity_labels,
        'begin_labels': begin_labels
    }

def train_entity_specific_auc_2t(
    data,
    target_entity_type,
    model_name=DEFAULT_MODEL_NAME,
    learning_rate=LEARNING_RATE,  # Use global variable
    num_epochs=NUM_EPOCHS,  # Use global variable
    batch_size=BATCH_SIZE,  # Use global variable
    lambda_param=LAMBDA_PARAM,  # Use global variable
    margin=MARGIN,  # Use global variable
    filter_training_data=False,  # New parameter to filter training data
    device='cuda' if torch.cuda.is_available() else 'cpu'
):
    """
    Train a NER model using the AUC-2T method, focusing on a specific entity type.
    
    Args:
        data: DataFrame containing NER data
        target_entity_type: The specific entity type to focus on (e.g., 'PERSON', 'LOCATION')
        model_name: Name of the pre-trained model to use
        learning_rate: Learning rate for the optimizer
        num_epochs: Number of training epochs
        batch_size: Batch size for training
        lambda_param: Trade-off parameter for combining the two AUC losses
        margin: Margin parameter for the Deep AUC Margin loss
        filter_training_data: If True, only include sentences with the target entity type in training
        device: Device to use for training ('cuda' or 'cpu')
        
    Returns:
        model: Trained NER model specialized for the target entity type
        history: Dictionary containing training metrics
    """
    tokenizer = AutoTokenizer.from_pretrained(model_name, local_files_only=_local_only_enabled())
    model = NERModelAUC2T(model_name=model_name).to(device)
    
    # Step 1: Data Preprocessing with entity type focus
    print(f"Preprocessing data for entity type: {target_entity_type}...")
    if filter_training_data:
        # Filter data to only include sentences with the target entity type
        data = data[data['raw_tags'].str.contains(f"B-{target_entity_type}", na=False)]
        print(f"Filtered training data to {len(data)} rows containing the entity type {target_entity_type}.")
    
    sentences, entity_labels, begin_labels = preprocess_entity_specific_data(data, tokenizer, target_entity_type)
    
    # Create dataset and dataloader
    dataset = NERDataset(sentences, entity_labels, begin_labels)
    dataloader = DataLoader(dataset, batch_size=batch_size, shuffle=True, collate_fn=custom_collate_fn)
    
    # Step 2: Initialize Model and Optimizer
    optimizer = optim.Adam(model.parameters(), lr=learning_rate)
    
    # Step 3: Training Loop
    print(f"Starting training for entity type {target_entity_type} for {num_epochs} epochs...")
    history = {
        'entity_loss': [],
        'begin_loss': [],
        'total_loss': []
    }
    
    for epoch in range(num_epochs):
        model.train()
        epoch_entity_loss = 0
        epoch_begin_loss = 0
        epoch_total_loss = 0
        
        progress_bar = tqdm(dataloader, desc=f"Epoch {epoch+1}/{num_epochs}")
        for batch in progress_bar:
            sentences = batch['sentence']
            entity_label_batch = batch['entity_labels']
            begin_label_batch = batch['begin_labels']
            
            # Tokenize the sentences
            encoding = tokenizer(sentences, padding=True, truncation=True, return_tensors="pt")
            input_ids = encoding['input_ids'].to(device)
            attention_mask = encoding['attention_mask'].to(device)
            
            # Forward pass
            optimizer.zero_grad()
            entity_scores, begin_scores = model(input_ids, attention_mask=attention_mask)
            
            # Prepare labels and calculate loss
            batch_entity_loss = 0
            batch_begin_loss = 0
            
            for i in range(len(sentences)):
                # Get token labels for current sentence
                entity_labels_tensor = torch.tensor(entity_label_batch[i], device=device)
                begin_labels_tensor = torch.tensor(begin_label_batch[i], device=device)
                
                # Get predictions for current sentence (up to the length of the labels)
                sent_len = min(len(entity_label_batch[i]), entity_scores.size(1))
                sent_entity_scores = entity_scores[i, :sent_len]
                sent_begin_scores = begin_scores[i, :sent_len]
                
                # Slice labels to match prediction length
                sent_entity_labels = entity_labels_tensor[:sent_len]
                sent_begin_labels = begin_labels_tensor[:sent_len]
                
                # Skip if no labels
                if sent_len == 0:
                    continue
                
                # Calculate mask for valid tokens (not special tokens like [CLS], [SEP], [PAD])
                valid_mask = torch.ones(sent_len, device=device)
                
                # Calculate losses for entity and beginning classification
                entity_loss = calculate_DAM_loss(sent_entity_scores, sent_entity_labels, margin, valid_mask)
                begin_loss = calculate_DAM_loss(sent_begin_scores, sent_begin_labels, margin, valid_mask)
                
                batch_entity_loss += entity_loss
                batch_begin_loss += begin_loss
            
            # Average the losses over the batch
            batch_size_actual = len(sentences)
            if batch_size_actual > 0:
                batch_entity_loss = batch_entity_loss / batch_size_actual
                batch_begin_loss = batch_begin_loss / batch_size_actual
            
            # Combine the losses
            total_loss = batch_entity_loss + lambda_param * batch_begin_loss
            
            # Backward pass and update
            total_loss.backward()
            optimizer.step()
            
            # Track metrics
            epoch_entity_loss += batch_entity_loss.item()
            epoch_begin_loss += batch_begin_loss.item()
            epoch_total_loss += total_loss.item()
            
            # Update progress bar
            progress_bar.set_postfix({
                'entity_loss': batch_entity_loss.item(),
                'begin_loss': batch_begin_loss.item(),
                'total_loss': total_loss.item()
            })
        
        # Average losses for the epoch
        num_batches = len(dataloader)
        epoch_entity_loss /= num_batches
        epoch_begin_loss /= num_batches
        epoch_total_loss /= num_batches
        
        # Store in history
        history['entity_loss'].append(epoch_entity_loss)
        history['begin_loss'].append(epoch_begin_loss)
        history['total_loss'].append(epoch_total_loss)
        
        print(f"Epoch {epoch+1}/{num_epochs} - "
              f"Entity Loss: {epoch_entity_loss:.4f}, "
              f"Begin Loss: {epoch_begin_loss:.4f}, "
              f"Total Loss: {epoch_total_loss:.4f}")
    
    print(f"Training for entity type {target_entity_type} completed!")
    
    return model, history

def preprocess_entity_specific_data(data, tokenizer, target_entity_type, connect_subwords=False, dbg=False):
    """
    Preprocess training data into binary classification format for a specific entity type.
    
    Args:
        data: DataFrame containing sentences and their original BIO labels
        tokenizer: Tokenizer for the model
        target_entity_type: The specific entity type to focus on (e.g., 'PERSON', 'LOCATION')
        connect_subwords: If True, merge subwords (starting with ##) into one word
        dbg: If True, print debug information and sample sentences
        
    Returns:
        sentences: List of sentences
        entity_labels: Binary labels for the entity-token task (1 for target entity, -1 for non-entity/other entity)
        begin_labels: Binary labels for the beginning-token task (1 for beginning of target entity, -1 for others)
    """
    # Check the actual column names in the dataframe
    columns = data.columns.tolist()
    
    # Determine which columns to use based on what's available
    sentence_col = 'Sentence #' if 'Sentence #' in columns else 'id'
    word_col = 'Word' if 'Word' in columns else 'token'
    
    # Determine tag column - prefer raw_tags if available, otherwise Tag or ner_tags
    if 'raw_tags' in columns:
        tag_col = 'raw_tags'
    elif 'Tag' in columns:
        tag_col = 'Tag'
    elif 'ner_tags' in columns:
        tag_col = 'ner_tags'
    else:
        raise ValueError("No tag column found in the data")
    
    if dbg:
        print(f"[DBG] Using columns: sentence_col={sentence_col}, word_col={word_col}, tag_col={tag_col}")
        print(f"[DBG] Target entity type: {target_entity_type}")
    
    # Group by Sentence # to get sentences and their token labels
    sentences = []
    all_entity_labels = []
    all_begin_labels = []
    
    # Group by sentence ID
    for sentence_id, group in data.groupby(sentence_col):
        if dbg and sentence_id <= 5: print(data.loc[data.id == sentence_id])
        words = group[word_col].tolist()
        tags = group[tag_col].tolist()
        
        # Handle subword merging if connect_subwords is True
        if connect_subwords:
            merged_words = []
            merged_tags = []
            current_word = ""
            current_tag = None
            
            for word, tag in zip(words, tags):
                word = str(word)  # Ensure word is a string
                if word.startswith("##"):
                    current_word += word[2:]  # Append subword without ##
                else:
                    if current_word:  # Add the previous merged word
                        merged_words.append(current_word)
                        merged_tags.append(current_tag)
                    current_word = word
                    current_tag = tag
            
            # Add the last merged word
            if current_word:
                merged_words.append(current_word)
                merged_tags.append(current_tag)
            
            words = merged_words
            tags = merged_tags
        
        if dbg and sentence_id <= 5:  # Debug first 5 sentences
            print(f"[DBG] Original sentence {sentence_id}: words={words}")
            print(f"[DBG] Original sentence {sentence_id}: tags={tags}")
        
        # Filter out only specific special tokens and truly empty tokens
        filtered_words = []
        filtered_tags = []
        
        for word, tag in zip(words, tags):
            if dbg and sentence_id <= 5: print(f"[DBG 0.1] id {sentence_id} Processing word: '{word}' with tag: '{tag}'")
            # Convert word to string first to handle any type issues
            word_str = str(word) if word is not None else ""
            if dbg and sentence_id <= 5: print(f"[DBG 0.2] id {sentence_id} Converted word to string: '{word_str} original word {word}'")

            # Skip if word is None or NaN
            if pd.isna(word_str):
                if dbg and sentence_id <= 5: print(f"[DBG 0.3] Skipping NaN word in sentence {sentence_id}")
                continue
            # Check if word should be filtered out
            should_filter = False
            
            # Filter out explicit special tokens
            if word_str in ['[CLS]', '[SEP]']:
                should_filter = True
                if dbg and sentence_id <= 5: print(f"[DBG 0.4] Filtering out special token in sentence {sentence_id}: {word_str} original word: {word}")
            # Filter out empty strings or whitespace-only strings
            elif len(word_str.strip()) == 0:
                should_filter = True
                if dbg and sentence_id <= 5: print(f"[DBG 0.5] Filtering out empty string in sentence {sentence_id}: {word_str} original word: {word}")
            # Filter out the string 'nan' (which comes from pd.isna conversion to string)
            elif word_str.lower() == 'nan':
                should_filter = True
                if dbg and sentence_id <= 5: print(f"[DBG 0.6] Filtering out 'nan' string in sentence {sentence_id}: {word_str} original word: {word}")
            if not should_filter:
                filtered_words.append(word_str.strip())
                filtered_tags.append(tag)
        
        if dbg and sentence_id <= 5:
            print(f"[DBG 1.1] Filtered sentence {sentence_id}: words={filtered_words}")
            print(f"[DBG 1.2] Filtered sentence {sentence_id}: tags={filtered_tags}")
        
        # Skip if no valid words remain
        if not filtered_words:
            if dbg:
                print(f"[DBG] Skipping empty sentence {sentence_id}")
            continue
        
        # Transform BIO labels to entity-specific binary labels
        entity_labels, begin_labels = transform_bio_labels_to_entity_specific(filtered_tags, target_entity_type)
        
        if dbg and sentence_id <= 5:
            print(f"[DBG] Entity labels for sentence {sentence_id}: {entity_labels}")
            print(f"[DBG] Begin labels for sentence {sentence_id}: {begin_labels}")
        
        # Create sentence by joining words with spaces
        sentence_text = ' '.join(filtered_words)
        sentences.append(sentence_text)
        all_entity_labels.append(entity_labels)
        all_begin_labels.append(begin_labels)
        
        if dbg and sentence_id <= 5:
            print(f"[DBG] Final sentence {sentence_id}: '{sentence_text}'")
            print("-" * 50)
    
    if dbg:
        print(f"[DBG] Total processed sentences: {len(sentences)}")
        print(f"[DBG] Sample sentences:")
        #ok
        
        for i, (sent, ent_labels, beg_labels) in enumerate(zip(sentences[:5], all_entity_labels[:5], all_begin_labels[:5])):
            # Add RLM mark to help console display Hebrew correctly
            rlm = "\u200F"
            print(f"  {i+1}: '{rlm}{sent}'")
            print(f"     Entity labels: {ent_labels}")
            print(f"     Begin labels:  {beg_labels}")
    
    return sentences, all_entity_labels, all_begin_labels

def transform_bio_labels_to_entity_specific(original_labels, target_entity_type):
    """
    Transform original BIO labels to binary labels for entity and beginning tokens,
    focusing only on the specified entity type.
    
    Args:
        original_labels: List of original BIO labels for each token (string)
        target_entity_type: The specific entity type to focus on (e.g., 'PERSON', 'LOCATION')
        
    Returns:
        entity_labels: Binary labels for the entity-token task 
                      (1 for target entity, -1 for non-entity or other entity types)
        begin_labels: Binary labels for the beginning-token task 
                     (1 for beginning of target entity, -1 for others)
    """
    entity_labels = []
    begin_labels = []
    
    for label in original_labels:
        # Skip special tokens
        if label in ["[CLS]", "[SEP]", "[PAD]"] or label is None:
            continue
            
        # Standardize the label format (could be "B-PERSON" or just numeric)
        if isinstance(label, str):
            # Default is not an entity of target type
            is_target_entity = False
            is_beginning = False
            
            # Check if the label matches our target entity
            if label.startswith("B-") and target_entity_type in label:
                is_target_entity = True
                is_beginning = True
            elif label.startswith("I-") and target_entity_type in label:
                is_target_entity = True
                is_beginning = False
        else:
            # For numeric labels, you would need a mapping to determine the entity type
            # Here we assume a mapping is available elsewhere
            is_target_entity = False
            is_beginning = False
        
        # Convert to binary labels
        entity_labels.append(1 if is_target_entity else -1)
        begin_labels.append(1 if is_beginning else -1)
            
    return entity_labels, begin_labels

def predict_entity_specific(model, sentences, tokenizer, target_entity_type, threshold=0.0, device='cuda' if torch.cuda.is_available() else 'cpu'):
    """
    Make predictions using the trained entity-specific AUC-2T model.
    
    Args:
        model: Trained entity-specific AUC-2T model
        sentences: List of sentences to predict
        tokenizer: Tokenizer for the model
        target_entity_type: The specific entity type the model was trained for
        threshold: Decision threshold for binary classification
        device: Device to use for inference
        
    Returns:
        entity_predictions: Binary predictions for entity tokens of the target type
        begin_predictions: Binary predictions for beginning tokens of the target type
        bio_predictions: BIO format predictions for the sentences
    """
    model.eval()
    entity_predictions = []
    begin_predictions = []
    bio_predictions = []
    
    with torch.no_grad():
        for sentence in sentences:
            # Tokenize the sentence
            encoding = tokenizer(sentence, padding=True, truncation=True, return_tensors="pt")
            input_ids = encoding['input_ids'].to(device)
            attention_mask = encoding['attention_mask'].to(device)
            
            # Get model predictions
            entity_scores, begin_scores = model(input_ids, attention_mask=attention_mask)
            
            # Convert scores to binary predictions based on threshold
            entity_preds = torch.sign(entity_scores - threshold).squeeze().cpu().numpy()
            begin_preds = torch.sign(begin_scores - threshold).squeeze().cpu().numpy()
            
            # Filter out special tokens (like [CLS], [SEP], [PAD])
            tokens = tokenizer.convert_ids_to_tokens(input_ids[0])
            filtered_entity_preds = []
            filtered_begin_preds = []
            
            for i, token in enumerate(tokens):
                if token not in ["[CLS]", "[SEP]", "[PAD]"]:
                    filtered_entity_preds.append(entity_preds[i])
                    filtered_begin_preds.append(begin_preds[i])
            
            entity_predictions.append(filtered_entity_preds)
            begin_predictions.append(filtered_begin_preds)
            
            # Convert binary predictions to BIO format with the specific entity type
            bio_tags = convert_binary_to_specific_bio(filtered_entity_preds, filtered_begin_preds, target_entity_type)
            bio_predictions.append(bio_tags)
    
    return entity_predictions, begin_predictions, bio_predictions

def convert_binary_to_specific_bio(entity_preds, begin_preds, entity_type):
    """
    Convert binary predictions back to BIO format with a specific entity type.
    
    Args:
        entity_preds: Binary predictions for entity tokens (-1 or 1)
        begin_preds: Binary predictions for beginning tokens (-1 or 1)
        entity_type: The specific entity type (e.g., 'PERSON', 'LOCATION')
        
    Returns:
        bio_tags: List of BIO tags with the specific entity type
    """
    bio_tags = []
    current_entity = False
    
    # Make sure we're working with Python lists, not numpy arrays
    if hasattr(entity_preds, 'tolist'):
        entity_preds = entity_preds.tolist()
    if hasattr(begin_preds, 'tolist'):
        begin_preds = begin_preds.tolist()
    
    for i in range(len(entity_preds)):
        is_entity = (entity_preds[i] == 1)
        is_begin = (begin_preds[i] == 1)
        
        if not is_entity:
            bio_tags.append("O")
            current_entity = False
        elif is_begin:
            bio_tags.append(f"B-{entity_type}")
            current_entity = True
        elif current_entity:
            bio_tags.append(f"I-{entity_type}")
        else:
            # This shouldn't happen in theory but we'll handle it by assuming it's a beginning
            bio_tags.append(f"B-{entity_type}")
            current_entity = True
    
    return bio_tags

def evaluate_entity_specific_model(model, test_data, tokenizer, target_entity_type, device='cuda' if torch.cuda.is_available() else 'cpu'):
    """
    Evaluate the entity-specific AUC-2T model on test data.
    
    Args:
        model: Trained entity-specific AUC-2T model
        test_data: DataFrame containing test data
        tokenizer: Tokenizer for the model
        target_entity_type: The specific entity type the model was trained for
        device: Device to use for inference
        
    Returns:
        metrics: Dictionary containing evaluation metrics for the specific entity type
    """
    # Import here to avoid requiring seqeval if not evaluating
    try:
        from seqeval.metrics import classification_report, accuracy_score, precision_score, recall_score, f1_score
    except ImportError:
        print("Warning: seqeval not installed, metrics will be limited")
        # Define dummy functions that return 0
        def accuracy_score(*args, **kwargs): return 0
        def precision_score(*args, **kwargs): return 0
        def recall_score(*args, **kwargs): return 0
        def f1_score(*args, **kwargs): return 0
        def classification_report(*args, **kwargs): return {}
    
    # Preprocess test data for the specific entity type
    sentences, true_entity_labels, true_begin_labels = preprocess_entity_specific_data(test_data, tokenizer, target_entity_type)
    
    # Get predictions
    pred_entity_labels, pred_begin_labels, pred_bio = predict_entity_specific(model, sentences, tokenizer, target_entity_type, device=device)
    
    # Reconstruct true BIO tags
    true_bio_tags = []
    for sentence_idx in range(len(sentences)):
        # Get true BIO tags for the specific entity
        entity_is_positive = true_entity_labels[sentence_idx]
        begin_is_positive = true_begin_labels[sentence_idx]
        
        # Reconstruct BIO tags with the specific entity type
        sentence_true_bio = []
        for e, b in zip(entity_is_positive, begin_is_positive):
            if e == -1:
                sentence_true_bio.append("O")
            elif b == 1:
                sentence_true_bio.append(f"B-{target_entity_type}")
            else:
                sentence_true_bio.append(f"I-{target_entity_type}")
        
        true_bio_tags.append(sentence_true_bio)
    
    # Ensure pred and true tags have the same length
    for sentence_idx in range(len(sentences)):
        min_len = min(len(true_bio_tags[sentence_idx]), len(pred_bio[sentence_idx]))
        true_bio_tags[sentence_idx] = true_bio_tags[sentence_idx][:min_len]
        pred_bio[sentence_idx] = pred_bio[sentence_idx][:min_len]
    
    # Calculate metrics using seqeval
    try:
        report = classification_report(true_bio_tags, pred_bio, output_dict=True)
    except Exception as e:
        print(f"Error calculating classification report: {e}")
        report = {}
    
    # Extract overall metrics
    try:
        metrics = {
            'accuracy': accuracy_score(true_bio_tags, pred_bio),
            'precision': precision_score(true_bio_tags, pred_bio),
            'recall': recall_score(true_bio_tags, pred_bio),
            'f1': f1_score(true_bio_tags, pred_bio),
            'report': report
        }
    except Exception as e:
        print(f"Error calculating metrics: {e}")
        metrics = {'error': str(e)}
    
    return metrics

def evaluate_entity_specific_model_detailed(model, test_data, tokenizer, target_entity_type, device='cuda' if torch.cuda.is_available() else 'cpu'):
    """
    Evaluate the entity-specific AUC-2T model on test data with detailed results.
    
    Args:
        model: Trained entity-specific AUC-2T model
        test_data: DataFrame containing test data
        tokenizer: Tokenizer for the model
        target_entity_type: The specific entity type the model was trained for
        device: Device to use for inference
        
    Returns:
        metrics: Dictionary containing evaluation metrics for the specific entity type
        detailed_results: Dictionary with predictions, true labels, and confusion matrix
    """
    # Import here to avoid requiring seqeval if not evaluating
    try:
        from seqeval.metrics import classification_report as seq_classification_report, accuracy_score, precision_score, recall_score, f1_score
    except ImportError:
        print("Warning: seqeval not installed, metrics will be limited")
        # Define dummy functions that return 0
        def accuracy_score(*args, **kwargs): return 0
        def precision_score(*args, **kwargs): return 0
        def recall_score(*args, **kwargs): return 0
        def f1_score(*args, **kwargs): return 0
        def seq_classification_report(*args, **kwargs): return {}
    
    # Preprocess test data for the specific entity type
    sentences, true_entity_labels, true_begin_labels = preprocess_entity_specific_data(test_data, tokenizer, target_entity_type, dbg=False)
    
    # Get original tags for reference
    original_tags_per_sentence = get_original_tags_by_sentence(test_data)
    
    # Get predictions
    pred_entity_labels, pred_begin_labels, pred_bio = predict_entity_specific(model, sentences, tokenizer, target_entity_type, device=device)
    
    # Reconstruct true BIO tags
    true_bio_tags = []
    for sentence_idx in range(len(sentences)):
        # Get true BIO tags for the specific entity
        entity_is_positive = true_entity_labels[sentence_idx]
        begin_is_positive = true_begin_labels[sentence_idx]
        
        # Reconstruct BIO tags with the specific entity type
        sentence_true_bio = []
        for e, b in zip(entity_is_positive, begin_is_positive):
            if e == -1:
                sentence_true_bio.append("O")
            elif b == 1:
                sentence_true_bio.append(f"B-{target_entity_type}")
            else:
                sentence_true_bio.append(f"I-{target_entity_type}")
        
        true_bio_tags.append(sentence_true_bio)
    
    # Ensure pred and true tags have the same length
    for sentence_idx in range(len(sentences)):
        min_len = min(len(true_bio_tags[sentence_idx]), len(pred_bio[sentence_idx]))
        true_bio_tags[sentence_idx] = true_bio_tags[sentence_idx][:min_len]
        pred_bio[sentence_idx] = pred_bio[sentence_idx][:min_len]
    
    # Flatten for confusion matrix and detailed analysis
    flat_true = [tag for sentence in true_bio_tags for tag in sentence]
    flat_pred = [tag for sentence in pred_bio for tag in sentence]
    
    # Create confusion matrix
    unique_labels = sorted(list(set(flat_true + flat_pred)))
    conf_matrix = confusion_matrix(flat_true, flat_pred, labels=unique_labels)
    
    # Calculate metrics using seqeval
    try:
        report = seq_classification_report(true_bio_tags, pred_bio, output_dict=True)
    except Exception as e:
        print(f"Error calculating classification report: {e}")
        report = {}
    
    # Extract overall metrics
    try:
        metrics = {
            'accuracy': accuracy_score(true_bio_tags, pred_bio),
            'precision': precision_score(true_bio_tags, pred_bio),
            'recall': recall_score(true_bio_tags, pred_bio),
            'f1': f1_score(true_bio_tags, pred_bio),
            'report': report
        }
    except Exception as e:
        print(f"Error calculating metrics: {e}")
        metrics = {'error': str(e)}
    
    # Prepare detailed results
    detailed_results = {
        'sentences': sentences,
        'true_bio_tags': true_bio_tags,
        'pred_bio_tags': pred_bio,
        'original_tags': original_tags_per_sentence,
        'flat_true': flat_true,
        'flat_pred': flat_pred,
        'confusion_matrix': conf_matrix,
        'unique_labels': unique_labels,
        'entity_type': target_entity_type
    }
    
    return metrics, detailed_results

def get_original_tags_by_sentence(data):
    """
    Extract original tags grouped by sentence from the dataset.
    
    Args:
        data: DataFrame containing the dataset
        
    Returns:
        original_tags_per_sentence: List of lists containing original tags for each sentence
    """
    # Check the actual column names in the dataframe
    columns = data.columns.tolist()
    
    # Determine which columns to use based on what's available
    sentence_col = 'Sentence #' if 'Sentence #' in columns else 'id'
    word_col = 'Word' if 'Word' in columns else 'token'
    
    # Determine tag column - prefer raw_tags if available, otherwise Tag or ner_tags
    if 'raw_tags' in columns:
        tag_col = 'raw_tags'
    elif 'Tag' in columns:
        tag_col = 'Tag'
    elif 'ner_tags' in columns:
        tag_col = 'ner_tags'
    else:
        raise ValueError("No tag column found in the data")
    
    original_tags_per_sentence = []
    
    # Group by sentence ID
    for sentence_id, group in data.groupby(sentence_col):
        words = group[word_col].tolist()
        tags = group[tag_col].tolist()
        
        # Filter out special tokens and empty tokens (same filtering as in preprocessing)
        filtered_tags = []
        
        for word, tag in zip(words, tags):
            # Convert word to string first to handle any type issues
            word_str = str(word) if word is not None else ""
            
            # Skip if word is None or NaN
            if pd.isna(word_str):
                continue
                
            # Check if word should be filtered out
            should_filter = False
            
            # Filter out explicit special tokens
            if word_str in ['[CLS]', '[SEP]']:
                should_filter = True
            # Filter out empty strings or whitespace-only strings
            elif len(word_str.strip()) == 0:
                should_filter = True
            # Filter out the string 'nan' (which comes from pd.isna conversion to string)
            elif word_str.lower() == 'nan':
                should_filter = True
                
            if not should_filter:
                filtered_tags.append(tag)
        
        # Skip if no valid tags remain
        if filtered_tags:
            original_tags_per_sentence.append(filtered_tags)
    
    return original_tags_per_sentence

def export_results_to_excel_separate(models_results, output_dir="results"):
    """
    Export each entity model results to separate Excel files.
    
    Args:
        models_results: Dictionary with entity_type -> (metrics, detailed_results)
        output_dir: Directory to save the Excel files
    """
    # Create output directory if it doesn't exist
    if not os.path.exists(output_dir):
        os.makedirs(output_dir)
    
    for entity_type, (metrics, detailed_results) in models_results.items():
        # Create output filename with entity type
        output_file = os.path.join(output_dir, f"ner_auc2t_results_{entity_type}.xlsx")
        
        # Create a new workbook for this entity
        wb = openpyxl.Workbook()
        
        # Remove the default sheet
        wb.remove(wb.active)
        
        # Create summary sheet
        summary_sheet = wb.create_sheet("Summary")
        
        # Add metrics summary
        summary_sheet.append(["Metric", "Value"])
        summary_sheet.append(["Entity Type", str(entity_type)])  # Convert entity_type to string
        summary_sheet.append(["Accuracy", f"{metrics.get('accuracy', 0):.4f}"])
        summary_sheet.append(["Precision", f"{metrics.get('precision', 0):.4f}"])
        summary_sheet.append(["Recall", f"{metrics.get('recall', 0):.4f}"])
        summary_sheet.append(["F1-Score", f"{metrics.get('f1', 0):.4f}"])
        summary_sheet.append([])  # Empty row
        
        # Add confusion matrix
        conf_matrix = detailed_results['confusion_matrix']
        unique_labels = detailed_results['unique_labels']
        
        summary_sheet.append(["Confusion Matrix"])
        
        # Add column headers (predicted labels)
        header_row = ["True\\Pred"] + unique_labels
        summary_sheet.append(header_row)
        
        # Add confusion matrix data
        for i, true_label in enumerate(unique_labels):
            row_data = [true_label] + [int(conf_matrix[i][j]) for j in range(len(unique_labels))]
            summary_sheet.append(row_data)
        
        # Create detailed predictions sheet
        detail_sheet = wb.create_sheet("Sentence_Predictions")
        
        # Add headers
        detail_sheet.append(["Sentence_ID", "Sentence", "Original_Tags", "True_Tags_for_Entity", "Predicted_Tags", "Correct"])
        
        # Add detailed predictions
        for i, (sentence, original_tags, true_tags, pred_tags) in enumerate(zip(
            detailed_results['sentences'],
            detailed_results['original_tags'],
            detailed_results['true_bio_tags'],
            detailed_results['pred_bio_tags']
        )):
            # Check if sentence prediction is completely correct
            is_correct = "✓" if true_tags == pred_tags else "✗"
            
            detail_sheet.append([
                i + 1,
                sentence,
                " ".join(original_tags),
                " ".join(true_tags),
                " ".join(pred_tags),
                is_correct
            ])
        
        # Create token-level comparison sheet
        token_sheet = wb.create_sheet("Token_Level_Analysis")
        token_sheet.append(["Sentence_ID", "Token_ID", "Token", "Original_Tag", "True_Tag_for_Entity", "Predicted_Tag", "Match"])

        # Build token-level mapping: sentence id, token id, token text
        # We need to reconstruct tokens per sentence
        for sent_idx, sentence in enumerate(detailed_results['sentences']):
            # Split sentence into tokens
            tokens = sentence.split()
            original_tags = detailed_results['original_tags'][sent_idx]
            true_tags = detailed_results['true_bio_tags'][sent_idx]
            pred_tags = detailed_results['pred_bio_tags'][sent_idx]
            
            # Use min length to avoid index errors
            min_len = min(len(tokens), len(original_tags), len(true_tags), len(pred_tags))
            for token_idx in range(min_len):
                token = tokens[token_idx]
                original_tag = original_tags[token_idx]
                true_tag = true_tags[token_idx]
                pred_tag = pred_tags[token_idx]
                match = "✓" if true_tag == pred_tag else "✗"
                token_sheet.append([sent_idx + 1, token_idx + 1, token, original_tag, true_tag, pred_tag, match])
        
        # Create classification report sheet if available
        if 'report' in metrics and metrics['report']:
            report_sheet = wb.create_sheet("Classification_Report")
            report_sheet.append(["Label", "Precision", "Recall", "F1-Score", "Support"])
            
            for label, values in metrics['report'].items():
                if isinstance(values, dict) and 'precision' in values:
                    report_sheet.append([
                        label,
                        f"{values.get('precision', 0):.4f}",
                        f"{values.get('recall', 0):.4f}",
                        f"{values.get('f1-score', 0):.4f}",
                        str(values.get('support', 0))
                    ])
        
        # Save the workbook
        wb.save(output_file)
        print(f"Results for {entity_type} exported to {output_file}")

def train_incremental_auc2t_subsets(
    data, 
    entity_types, 
    subset_sizes=[50, 100, 150, 200, 250], 
    model_name='dicta-il/dictabert',
    learning_rate=LEARNING_RATE,  # Use global variable
    num_epochs=NUM_EPOCHS,  # Use global variable
    batch_size=BATCH_SIZE,  # Use global variable
    lambda_param=LAMBDA_PARAM,  # Use global variable
    margin=MARGIN,  # Use global variable
    filter_training_data=False,
    device='cuda' if torch.cuda.is_available() else 'cpu'
):
    """
    Train AUC-2T models on incrementally increasing random subsets of sentences for each entity type.
    
    Args:
        data: The full dataset
        entity_types: List of entity types to train models for
        subset_sizes: List of subset sizes to train on
        model_name: The model to use for training
        learning_rate: Learning rate for training
        num_epochs: Number of epochs for training
        batch_size: Batch size for training
        lambda_param: Lambda parameter for AUC-2T loss
        margin: Margin parameter for DAM loss
        filter_training_data: Whether to filter training data to only sentences with target entity
        device: Device to use for training
    
    Returns:
        results_dict: Dictionary with entity_type -> DataFrame of performance metrics for each subset size
    """
    from sklearn.model_selection import train_test_split
    from collections import Counter
    
    # Split data into train and test at sentence level
    sentences = []
    current_sentence = []
    current_sentence_id = None

    for idx, row in data.iterrows():
        sentence_col = 'Sentence #' if 'Sentence #' in data.columns else 'id'
        if current_sentence_id != row[sentence_col]:
            if current_sentence:
                sentences.append(pd.DataFrame(current_sentence))
            current_sentence = []
            current_sentence_id = row[sentence_col]
        current_sentence.append(row)
    
    # Add the last sentence if exists
    if current_sentence:
        sentences.append(pd.DataFrame(current_sentence))

    # Split at sentence level for consistent test set
    train_sentences, test_sentences = train_test_split(sentences, test_size=0.2, random_state=42)
    test_data = pd.concat(test_sentences, ignore_index=True)
    
    print(f"Total sentences: {len(sentences)}")
    print(f"Training sentences available: {len(train_sentences)}")
    print(f"Test sentences: {len(test_sentences)}")
    
    results_dict = {}
    
    for entity_type in entity_types:
        print(f"\n{'='*60}")
        print(f"Processing entity type: {entity_type}")
        print(f"{'='*60}")
        
        # Filter train sentences based on entity type if requested
        if filter_training_data:
            # Only keep sentences that contain the target entity type
            filtered_train_sentences = []
            for sent_df in train_sentences:
                tag_col = 'raw_tags' if 'raw_tags' in sent_df.columns else 'Tag'
                if sent_df[tag_col].str.contains(f"B-{entity_type}", na=False).any():
                    filtered_train_sentences.append(sent_df)
            available_sentences = filtered_train_sentences
            print(f"Filtered to {len(available_sentences)} sentences containing {entity_type}")
        else:
            available_sentences = train_sentences
            print(f"Using all {len(available_sentences)} training sentences")
        
        # Filter subset sizes based on available sentences
        max_sentences = len(available_sentences)
        valid_subset_sizes = [size for size in subset_sizes if size <= max_sentences]
        
        if not valid_subset_sizes:
            print(f"No valid subset sizes for {entity_type}. Maximum available sentences: {max_sentences}")
            results_dict[entity_type] = pd.DataFrame()
            continue
        
        print(f"Training on subset sizes: {valid_subset_sizes} (max available: {max_sentences})")
        
        results = []
        
        for subset_size in valid_subset_sizes:
            print(f"\nTraining {entity_type} model on {subset_size} sentences...")
            
            # Randomly sample sentences for training
            random.seed(42)  # For reproducibility
            train_subset_sentences = random.sample(available_sentences, subset_size)
            train_subset_data = pd.concat(train_subset_sentences, ignore_index=True)
            
            try:
                # Train the model
                model, history = train_entity_specific_auc_2t(
                    train_subset_data,
                    target_entity_type=entity_type,
                    model_name=model_name,
                    learning_rate=learning_rate,
                    num_epochs=num_epochs,
                    batch_size=batch_size,
                    lambda_param=lambda_param,
                    margin=margin,
                    filter_training_data=False,  # Already filtered above if needed
                    device=device
                )
                
                # Evaluate the model
                tokenizer = AutoTokenizer.from_pretrained(model_name, local_files_only=_local_only_enabled())
                metrics, detailed_results = evaluate_entity_specific_model_detailed(
                    model, test_data, tokenizer, entity_type, device=device
                )
                
                # Extract metrics
                accuracy = metrics.get('accuracy', 0)
                precision = metrics.get('precision', 0)
                recall = metrics.get('recall', 0)
                f1_score = metrics.get('f1', 0)
                
                # Get final training losses
                final_entity_loss = history['entity_loss'][-1] if history['entity_loss'] else 0
                final_begin_loss = history['begin_loss'][-1] if history['begin_loss'] else 0
                final_total_loss = history['total_loss'][-1] if history['total_loss'] else 0
                
                # Calculate entity-specific metrics from classification report
                entity_precision = 0
                entity_recall = 0
                entity_f1 = 0
                entity_support = 0
                
                if 'report' in metrics and metrics['report']:
                    entity_key = f"B-{entity_type}"
                    if entity_key in metrics['report']:
                        entity_precision = metrics['report'][entity_key].get('precision', 0)
                        entity_recall = metrics['report'][entity_key].get('recall', 0)
                        entity_f1 = metrics['report'][entity_key].get('f1-score', 0)
                        entity_support = metrics['report'][entity_key].get('support', 0)
                
                # Store results
                result_row = {
                    'entity_type': entity_type,
                    'subset_size': subset_size,
                    'overall_accuracy': accuracy,
                    'overall_precision': precision,
                    'overall_recall': recall,
                    'overall_f1': f1_score,
                    'entity_precision': entity_precision,
                    'entity_recall': entity_recall,
                    'entity_f1': entity_f1,
                    'entity_support': entity_support,
                    'final_entity_loss': final_entity_loss,
                    'final_begin_loss': final_begin_loss,
                    'final_total_loss': final_total_loss,
                    'training_sentences': subset_size,
                    'filtered_training': filter_training_data
                }
                results.append(result_row)
                
                print(f"Results for {entity_type} with {subset_size} sentences:")
                print(f"  Overall - F1: {f1_score:.4f}, Precision: {precision:.4f}, Recall: {recall:.4f}")
                print(f"  Entity-specific - F1: {entity_f1:.4f}, Precision: {entity_precision:.4f}, Recall: {entity_recall:.4f}")
                
            except Exception as e:
                print(f"Error training {entity_type} model with {subset_size} sentences: {str(e)}")
                # Store error result
                result_row = {
                    'entity_type': entity_type,
                    'subset_size': subset_size,
                    'overall_accuracy': 0,
                    'overall_precision': 0,
                    'overall_recall': 0,
                    'overall_f1': 0,
                    'entity_precision': 0,
                    'entity_recall': 0,
                    'entity_f1': 0,
                    'entity_support': 0,
                    'final_entity_loss': 0,
                    'final_begin_loss': 0,
                    'final_total_loss': 0,
                    'training_sentences': subset_size,
                    'filtered_training': filter_training_data,
                    'error': str(e)
                }
                results.append(result_row)
        
        results_df = pd.DataFrame(results)
        results_dict[entity_type] = results_df
        
        print(f"\nCompleted incremental training for {entity_type}")
        if not results_df.empty:
            print("Summary of results:")
            print(results_df[['subset_size', 'overall_f1', 'entity_f1']].to_string(index=False))
    
    return results_dict

def save_incremental_auc2t_results(results_dict, output_dir="auc2t_incremental_results"):
    """
    Save the incremental AUC-2T training results to Excel files.
    
    Args:
        results_dict: Dictionary with entity_type -> DataFrame of results
        output_dir: Directory to save the Excel files
    """
    # Create output directory if it doesn't exist
    if not os.path.exists(output_dir):
        os.makedirs(output_dir)
    
    # Save individual entity results
    for entity_type, results_df in results_dict.items():
        if not results_df.empty:
            entity_output = os.path.join(output_dir, f"incremental_auc2t_{entity_type}.xlsx")
            with pd.ExcelWriter(entity_output, engine='openpyxl') as writer:
                results_df.to_excel(writer, sheet_name="IncrementalResults", index=False)
            print(f"Incremental AUC-2T results for {entity_type} saved to {entity_output}")
    
    # Save combined results
    if results_dict:
        combined_df = pd.concat(results_dict.values(), ignore_index=True)
        combined_output = os.path.join(output_dir, "incremental_auc2t_combined.xlsx")
        
        with pd.ExcelWriter(combined_output, engine='openpyxl') as writer:
            # Save combined results
            combined_df.to_excel(writer, sheet_name="AllResults", index=False)
            
            # Save summary by entity type
            if not combined_df.empty:
                summary_df = combined_df.groupby(['entity_type', 'subset_size']).agg({
                    'overall_f1': 'mean',
                    'entity_f1': 'mean',
                    'overall_precision': 'mean',
                    'overall_recall': 'mean',
                    'entity_precision': 'mean',
                    'entity_recall': 'mean'
                }).reset_index()
                summary_df.to_excel(writer, sheet_name="Summary", index=False)
                
                # Create pivot table for easy comparison
                pivot_overall_f1 = combined_df.pivot(index='subset_size', columns='entity_type', values='overall_f1')
                pivot_entity_f1 = combined_df.pivot(index='subset_size', columns='entity_type', values='entity_f1')
                
                pivot_overall_f1.to_excel(writer, sheet_name="OverallF1_Pivot")
                pivot_entity_f1.to_excel(writer, sheet_name="EntityF1_Pivot")
        
        print(f"Combined incremental AUC-2T results saved to {combined_output}")

def run_incremental_auc2t_experiment(
    file_path="ner_dataset.csv",
    subset_sizes=[50, 100, 150, 200, 250],
    model_name='dicta-il/dictabert',
    learning_rate=LEARNING_RATE,  # Use global variable
    num_epochs=NUM_EPOCHS,  # Use global variable
    batch_size=BATCH_SIZE,  # Use global variable
    lambda_param=LAMBDA_PARAM,  # Use global variable
    margin=MARGIN,  # Use global variable
    filter_training_data=True,
    min_entity_count=50,
    device='cuda' if torch.cuda.is_available() else 'cpu'
):
    """
    Run the complete incremental training experiment for AUC-2T models.
    
    Args:
        file_path: Path to the NER dataset CSV file
        subset_sizes: List of subset sizes to train on
        model_name: Pre-trained model to use
        learning_rate: Learning rate for training
        num_epochs: Number of epochs for training
        batch_size: Batch size for training
        lambda_param: Lambda parameter for AUC-2T loss
        margin: Margin parameter for DAM loss
        filter_training_data: Whether to filter training data to only sentences with target entity
        min_entity_count: Minimum count required for an entity type to be included
        device: Device to use for training
    
    Returns:
        results_dict: Dictionary with entity_type -> DataFrame of performance metrics
    """
    # Load data
    with open(file_path, 'rb') as f:
        result = chardet.detect(f.read())
        data = pd.read_csv(file_path, delimiter=',', encoding=result['encoding'])
    
    print(f"Loaded data with {len(data)} rows")
    
    # Get entity types and filter by count
    if 'raw_tags' in data.columns:
        entity_tags = [tag.split('-')[-1] for tag in data['raw_tags'] if isinstance(tag, str) and '-' in tag]
    elif 'Tag' in data.columns:
        entity_tags = [tag.split('-')[-1] for tag in data['Tag'] if isinstance(tag, str) and '-' in tag]
    else:
        raise ValueError("No suitable tag column found in dataset")

    # Count entity occurrences
    from collections import Counter
    entity_counts = Counter(entity_tags)
    print("Entity counts in data:")
    for entity, count in entity_counts.most_common():
        print(f"{entity}: {count}")

    # Filter entities with sufficient count
    filtered_entity_types = [entity for entity, count in entity_counts.items() if count >= min_entity_count]
    print(f"\nEntity types with count >= {min_entity_count}: {filtered_entity_types}")
    
    if not filtered_entity_types:
        print("No entity types meet the minimum count requirement!")
        return {}
    
    # Run incremental training
    print(f"\nRunning incremental training for {len(filtered_entity_types)} entity types...")
    results_dict = train_incremental_auc2t_subsets(
        data, 
        filtered_entity_types, 
        subset_sizes=subset_sizes, 
        model_name=model_name,
        learning_rate=learning_rate,
        num_epochs=num_epochs,
        batch_size=batch_size,
        lambda_param=lambda_param,
        margin=margin,
        filter_training_data=filter_training_data,
        device=device
    )
    
    # Save results
    output_dir = os.path.join(os.path.dirname(file_path), "auc2t_incremental_results")
    save_incremental_auc2t_results(results_dict, output_dir)
    
    return results_dict

# Example usage of entity-specific training
if __name__ == "__main__":
    # Load data
    file_path = "ner_dataset.csv"
    with open(file_path, 'rb') as f:
        result = chardet.detect(f.read())
        data = pd.read_csv(file_path, delimiter=',', encoding=result['encoding'])

    print(f"Loaded data with {len(data)} rows")

    # Print entity distribution
    if 'raw_tags' in data.columns:
        print("\nEntity distribution:")
        print(data['raw_tags'].value_counts())

    # Split data into train and test
    from sklearn.model_selection import train_test_split
    from collections import Counter

    # Group data into sentences first - using 'id' instead of 'Sentence #'
    sentences = []
    current_sentence = []
    current_sentence_id = None

    for idx, row in data.iterrows():
        if current_sentence_id != row['id']:
            if current_sentence:
                sentences.append(pd.DataFrame(current_sentence))
            current_sentence = []
            current_sentence_id = row['id']
        current_sentence.append(row)

    # Add the last sentence if exists
    if current_sentence:
        sentences.append(pd.DataFrame(current_sentence))

    # Now split at sentence level
    train_sentences, test_sentences = train_test_split(sentences, test_size=0.2, random_state=42)

    # Concatenate back into dataframes
    train_data = pd.concat(train_sentences, ignore_index=True)
    test_data = pd.concat(test_sentences, ignore_index=True)

    # Print split statistics
    print(f"Total sentences: {len(sentences)}")
    print(f"Training sentences: {len(train_sentences)}")
    print(f"Test sentences: {len(test_sentences)}")

    # Get the list of unique entity tags from the training data
    if 'raw_tags' in train_data.columns:
        tag_list = sorted(set(tag.split('-')[-1] for tag in train_data['raw_tags'] if isinstance(tag, str) and '-' in tag))
    elif 'Tag' in train_data.columns:
        tag_list = sorted(set(tag.split('-')[-1] for tag in train_data['Tag'] if isinstance(tag, str) and '-' in tag))
    elif 'ner_tags' in train_data.columns and 'raw_tags' in train_data.columns:
        tag_list = sorted(set(tag.split('-')[-1] for tag in train_data['raw_tags'] if isinstance(tag, str) and '-' in tag))
    else:
        tag_list = []
    print(f"Entity types in training data: {tag_list}")

    # Get all entity tags from raw_tags column
    if 'raw_tags' in train_data.columns:
        entity_tags = [tag.split('-')[-1] for tag in train_data['raw_tags'] if isinstance(tag, str) and '-' in tag]
    elif 'Tag' in train_data.columns:
        entity_tags = [tag.split('-')[-1] for tag in train_data['Tag'] if isinstance(tag, str) and '-' in tag]
    elif 'ner_tags' in train_data.columns and 'raw_tags' in train_data.columns:
        entity_tags = [tag.split('-')[-1] for tag in train_data['raw_tags'] if isinstance(tag, str) and '-' in tag]
    else:
        entity_tags = []

    # Count entity occurrences
    entity_counts = Counter(entity_tags)
    print("Entity counts in training data:")
    for entity, count in entity_counts.items():
        print(f"{entity}: {count}")

    # Only keep entities with count over N=50
    N = 50
    filtered_tag_list = [entity for entity in tag_list if entity_counts[entity] > N]
    print(f"Entity types with count > {N}: {filtered_tag_list}")
    tag_list = filtered_tag_list
    # print("Stopping here for testing. Exiting program.")
    # sys.exit(0)
    # Loop through all entity types and train a model for each
    models = {}
    histories = {}
    metrics_dict = {}

    # for entity_type in filtered_tag_list:
    #     print(f"\nTraining model for entity type: {entity_type}")
    #     model, history = train_entity_specific_auc_2t(
    #         train_data,
    #         target_entity_type=entity_type,
    #         model_name='dicta-il/dictabert',
    #         learning_rate=LEARNING_RATE,  # Use global variable
    #         num_epochs=NUM_EPOCHS,  # Use global variable
    #         batch_size=BATCH_SIZE,  # Use global variable
    #         lambda_param=LAMBDA_PARAM,  # Use global variable
    #         margin=MARGIN  # Use global variable
    #     )
    #     tokenizer = AutoTokenizer.from_pretrained('dicta-il/dictabert')

    #     # Get detailed evaluation results
    #     metrics, detailed_results = evaluate_entity_specific_model_detailed(model, test_data, tokenizer, entity_type)

    #     print(f"Evaluation metrics for {entity_type}:")
    #     print(f"Accuracy: {metrics['accuracy']:.4f}")
    #     print(f"Precision: {metrics['precision']:.4f}")
    #     print(f"Recall: {metrics['recall']:.4f}")
    #     print(f"F1 Score: {metrics['f1']:.4f}")

    #     models[entity_type] = model
    #     histories[entity_type] = history
    #     metrics_dict[entity_type] = (metrics, detailed_results)

    # # Export results to separate files for each entity
    # output_dir = os.path.join(os.path.dirname(__file__), "ner_results")
    # export_results_to_excel_separate(metrics_dict, output_dir)
    # for entity_type in filtered_tag_list:
    #     for filter_option in [True, False]:
    #         print(f"\nTraining with filter_training_data={filter_option} for entity type: {entity_type}")
    #         model, history = train_entity_specific_auc_2t(
    #             train_data,
    #             target_entity_type=entity_type,
    #             model_name='dicta-il/dictabert',
    #             learning_rate=LEARNING_RATE,  # Use global variable
    #             num_epochs=NUM_EPOCHS,  # Use global variable
    #             batch_size=BATCH_SIZE,  # Use global variable
    #             lambda_param=LAMBDA_PARAM,  # Use global variable
    #             margin=MARGIN,  # Use global variable
    #             filter_training_data=filter_option
    #         )
    #         tokenizer = AutoTokenizer.from_pretrained('dicta-il/dictabert')
            
    #         # Get detailed evaluation results
    #         metrics, detailed_results = evaluate_entity_specific_model_detailed(model, test_data, tokenizer, entity_type)
            
    #         print(f"Evaluation metrics for {entity_type} with filter_training_data={filter_option}:")
    #         print(f"Accuracy: {metrics['accuracy']:.4f}")
    #         print(f"Precision: {metrics['precision']:.4f}")
    #         print(f"Recall: {metrics['recall']:.4f}")
    #         print(f"F1 Score: {metrics['f1']:.4f}")
            
    #         models[(entity_type, filter_option)] = model
    #         histories[(entity_type, filter_option)] = history
    #         metrics_dict[(entity_type, filter_option)] = (metrics, detailed_results)
            
    #         # Export results to separate files for each entity
    #         output_dir = os.path.join(os.path.dirname(__file__), f"ner_results{'_filtered' if filter_option else ''}")
    #         export_results_to_excel_separate(metrics_dict, output_dir)

    # Incremental training on subsets
    print("\n" + "="*80)
    print("RUNNING INCREMENTAL TRAINING ON SUBSETS")
    print("="*80)
    
    subset_sizes = [50, 100, 150, 200, 250]
    results_dict = train_incremental_auc2t_subsets(
        data, 
        filtered_tag_list, 
        subset_sizes=subset_sizes, 
        model_name='dicta-il/dictabert',
        learning_rate=LEARNING_RATE,  # Use global variable
        num_epochs=NUM_EPOCHS,  # Use global variable
        batch_size=BATCH_SIZE,  # Use global variable
        lambda_param=LAMBDA_PARAM,  # Use global variable
        margin=MARGIN,  # Use global variable
        filter_training_data=True,
        device='cuda' if torch.cuda.is_available() else 'cpu'
    )
    
    # Save incremental results
    output_dir = os.path.join(os.path.dirname(__file__), "auc2t_incremental_results")
    save_incremental_auc2t_results(results_dict, output_dir)

    # Also run the standalone incremental experiment for easy execution
    print("\n" + "="*80)
    print("RUNNING STANDALONE INCREMENTAL EXPERIMENT")
    print("="*80)
    
    try:
        incremental_results = run_incremental_auc2t_experiment(
            file_path=file_path,
            subset_sizes=[50, 100, 150, 200, 250],
            model_name='dicta-il/dictabert',
            learning_rate=LEARNING_RATE,  # Use global variable
            num_epochs=NUM_EPOCHS,  # Use global variable
            batch_size=BATCH_SIZE,  # Use global variable
            lambda_param=LAMBDA_PARAM,  # Use global variable
            margin=MARGIN,  # Use global variable
            filter_training_data=True,
            min_entity_count=50,
            device='cuda' if torch.cuda.is_available() else 'cpu'
        )
        print(f"Incremental experiment completed for {len(incremental_results)} entity types")
    except Exception as e:
        print(f"Error in incremental experiment: {e}")

    # Run the complete incremental AUC-2T training experiment
    # Uncomment the following line to execute
    # run_incremental_auc2t_experiment(file_path, subset_sizes, 'dicta-il/dictabert', 2e-5, 3, 16, 100.0, 1.0, True, 50, 'cuda' if torch.cuda.is_available() else 'cpu')
