"""
merge_error_analysis_excels.py — Combine many per-model error-analysis Excel files
into a single workbook.

All the input files share the same template (same sheet names / columns) but each
represents a different model, split condition, seed, etc.

Merge rules
-----------
* Every sheet EXCEPT ``confusion_matrix`` is a simple vertical append: the header row
  is written once, then the data rows from every file are stacked underneath. A
  leading ``source_file`` column is added so each row can be traced back to its origin.
* The ``confusion_matrix`` sheet is special: matrices from different models can have
  different label sets, so they are NOT appended column-wise. Instead they are stacked
  one after another in a single sheet, each preceded by a bold title identifying the
  source (model / split_condition / seed / file name), with a blank row between them.

Usage
-----
    python merge_error_analysis_excels.py
    python merge_error_analysis_excels.py --input-dir "outputs/00 final_results/error analysis"
    python merge_error_analysis_excels.py --output combined.xlsx
"""
from __future__ import annotations

import argparse
import os
from pathlib import Path

import openpyxl
from openpyxl.styles import Font, PatternFill

CONFUSION_SHEET = "confusion_matrix"

# Default location of the per-model error-analysis Excel files.
_DEFAULT_INPUT_DIR = os.path.join("outputs", "00 final_results", "error analysis")


def _rows(ws) -> list[list]:
    """Return every row of a worksheet as a list of cell values."""
    return [[c.value for c in row] for row in ws.iter_rows()]


def _identity(wb, file_name: str) -> str:
    """Build a human-readable title for a file from its ``metrics`` sheet."""
    try:
        ws = wb["metrics"]
    except KeyError:
        return file_name

    header = [c.value for c in ws[1]]
    first_data = next(
        (row for row in ws.iter_rows(min_row=2, values_only=True) if any(v is not None for v in row)),
        None,
    )
    if not first_data:
        return file_name

    lookup = {str(h): v for h, v in zip(header, first_data)}
    parts = [
        str(lookup[key])
        for key in ("model", "split_condition", "seed")
        if lookup.get(key) not in (None, "")
    ]
    label = " | ".join(parts)
    return f"{label}  ({file_name})" if label else file_name


def merge(input_dir: str, output_path: str) -> None:
    output_stem = Path(output_path).stem  # e.g. "combined_error_analysis"
    files = sorted(
        p for p in Path(input_dir).glob("*.xlsx")
        # Skip Excel lock files (~$...) and the output itself, including any of its
        # copies/variants like "combined_error_analysis - Copy.xlsx".
        if not p.name.startswith("~$") and not p.stem.startswith(output_stem)
    )
    if not files:
        raise SystemExit(f"No .xlsx files found in: {input_dir}")

    out_wb = openpyxl.Workbook()
    out_wb.remove(out_wb.active)  # drop the default empty sheet

    # Track which normal sheets already have their header written.
    header_written: set[str] = set()
    title_font = Font(bold=True, size=12)
    title_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")

    # The confusion sheet is created lazily so it keeps a sensible position.
    confusion_ws = None
    confusion_row = 1

    skipped: list[str] = []
    merged = 0
    for path in files:
        try:
            wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
        except (PermissionError, OSError) as exc:
            print(f"[skip] could not open {path.name}: {exc} (is it open in Excel?)")
            skipped.append(path.name)
            continue
        merged += 1
        title = _identity(wb, path.name)

        for sheet_name in wb.sheetnames:
            src = wb[sheet_name]
            data = _rows(src)
            if not data:
                continue

            if sheet_name == CONFUSION_SHEET:
                if confusion_ws is None:
                    confusion_ws = out_wb.create_sheet(CONFUSION_SHEET)

                # Title row spanning the matrix width.
                title_cell = confusion_ws.cell(row=confusion_row, column=1, value=title)
                title_cell.font = title_font
                title_cell.fill = title_fill
                confusion_row += 1

                for row_values in data:
                    for col_idx, value in enumerate(row_values, start=1):
                        confusion_ws.cell(row=confusion_row, column=col_idx, value=value)
                    confusion_row += 1

                confusion_row += 1  # blank spacer row between matrices
                continue

            # Normal sheet: simple vertical append with a source_file column.
            dst = out_wb[sheet_name] if sheet_name in out_wb.sheetnames else out_wb.create_sheet(sheet_name)

            header, *body = data
            if sheet_name not in header_written:
                dst.append(["source_file", *[str(h) if h is not None else "" for h in header]])
                dst["A1"].font = Font(bold=True)
                header_written.add(sheet_name)

            for row_values in body:
                if all(v is None for v in row_values):
                    continue
                dst.append([path.name, *row_values])

        wb.close()

    try:
        out_wb.save(output_path)
    except PermissionError:
        raise SystemExit(
            f"[error] Cannot write '{output_path}' — it is open in Excel. "
            "Close the workbook and run this script again."
        )
    print(f"[done] merged {merged} of {len(files)} file(s) -> {output_path}")
    if skipped:
        print(f"[skipped] {len(skipped)} file(s): {skipped}")
    print(f"[sheets] {out_wb.sheetnames}")


def main() -> None:
    parser = argparse.ArgumentParser(description="Merge per-model error-analysis Excel files into one workbook.")
    parser.add_argument("--input-dir", default=_DEFAULT_INPUT_DIR,
                        help=f"Folder containing the .xlsx files (default: {_DEFAULT_INPUT_DIR}).")
    parser.add_argument("--output", default=None,
                        help="Output .xlsx path (default: <input-dir>/combined_error_analysis.xlsx).")
    args = parser.parse_args()

    output_path = args.output or os.path.join(args.input_dir, "combined_error_analysis.xlsx")
    merge(args.input_dir, output_path)


if __name__ == "__main__":
    main()
